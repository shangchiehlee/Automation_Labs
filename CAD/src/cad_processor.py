#!/usr/bin/env python3
"""cad_processor.py

Program Overview: CAD Contract Processor (Deterministic + Fuzzy Banding)
Author: Shane Lee

System Function:
    Transforms a CAD export workbook into a deterministic, auditable output workbook with
    processing metadata, trend analysis, a subject-level report, and fuzzy band assignments.

Architectural Pattern:
    Implements a deterministic batch pipeline with explicit data-quality accounting.

    1. Ingestion: Locates the header row in the source workbook and streams rows in read-only
       mode to avoid loading the entire sheet into memory.
    2. Processing: Aggregates subject-year and school-year metrics, computes cost-per-student,
       and derives fuzzy anchors (min/median/max) per year.
    3. Output: Writes multiple sheets (Processing Summary, Trend Analysis, Report, Fuzzy Bands)
       with formatting and audit metadata.

Purpose
- Convert a CAD export workbook (CAD_Contract.xlsx) into a processed workbook with:
  1) Processing_Summary: run metadata and observed data-quality counts
  2) Trend_Analysis: School-by-year cost-per-student view with conditional formatting
  3) Report: subject-level table used to derive School-Year aggregates
  4) Fuzzy_Bands: a small fuzzy-logic component that assigns Low/Medium/High bands per year

Evidence basis
- Statements about behaviour are supported only by the code and the produced workbooks.
- This script is designed for operational reproducibility. It does not claim scale properties
  beyond what can be demonstrated from runs and the implemented algorithmic discipline.

Inputs
- inputs/CAD_Contract.xlsx (Excel workbook). The script detects the main data table by
  locating a header row containing required fields (for example, 'School').

Outputs
- outputs/Processed_CAD_Contract.xlsx (Excel workbook) containing the sheets listed above.
- outputs/cad_processor.log and outputs/cad_processor_execution.log (best-effort).

Determinism and auditability
- Given identical input files and the same runtime environment, the transformation is intended
  to be deterministic. The Processing_Summary sheet records the input file SHA-256 to provide
  a stable fingerprint of the exact input used for a run.

Resource discipline
- Uses read-only worksheet iteration for header detection and schema inspection.
- Aggregation is performed with bounded intermediate structures (keys and summary rows).
  Peak memory depends on dataframe size and the number of unique groups, not solely on row count.

Fuzzy component
- The fuzzy banding uses a triangular membership function with per-year anchors (min, median, max)
  computed from finite, positive School-Year values. Membership values are decision-support signals,
  not probabilities.

Limitations
- Input schema assumptions are intentionally narrow and should be treated as data contracts.
  If headers or required columns change, header detection and column mapping may fail.
"""

from __future__ import annotations

import argparse
import datetime
import hashlib
import logging
import math
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# -------------------------
# Constants and schema
# -------------------------

RE_YEAR = re.compile(r"(19|20)\d{2}")
RE_SUMMARY_ROW = re.compile(r"^(total|sum|result)$", re.IGNORECASE)

DEFAULT_INPUT_FILENAME = "CAD_Contract.xlsx"
DEFAULT_OUTPUT_FILENAME = "Processed_CAD_Contract.xlsx"

# Canonical required columns
COL_SCHOOL = "School"
COL_SUBJECT_NO = "Subject No."
COL_SUBJECT = "Subject"
COL_TEACHING_SESSION = "Teaching Session"
COL_COST = "Incl Oncosts"
COL_STUDENT_COUNT = "Student Count"

REQUIRED_COLUMNS = [COL_SCHOOL, COL_SUBJECT_NO, COL_SUBJECT, COL_TEACHING_SESSION, COL_COST, COL_STUDENT_COUNT]

# Header normalisation: allow minor variants found in exports
COLUMN_ALIASES = {
    "subject no": COL_SUBJECT_NO,
    "subject no.": COL_SUBJECT_NO,
    "subject number": COL_SUBJECT_NO,
    "teaching session": COL_TEACHING_SESSION,
    "incl oncosts": COL_COST,
    "incl. oncosts": COL_COST,
    "student count": COL_STUDENT_COUNT,
}


# -------------------------
# Paths and logging
# -------------------------

@dataclass(frozen=True)
class CADPaths:
    base_dir: Path
    inputs_dir: Path
    outputs_dir: Path
    input_path: Path
    output_path: Path
    log_path: Path
    execution_log_path: Path


def _normalise_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _find_project_root(start_dir: Path, max_levels: int = 6) -> Path:
    """
    Walk up parent directories looking for a CAD project root.

    Heuristics:
    - A root contains config.json AND an inputs/ directory.
    - If not found, fall back to start_dir.
    """
    cur = start_dir.resolve()
    for _ in range(max_levels + 1):
        if (cur / "config.json").exists() and (cur / "inputs").is_dir():
            return cur
        if cur.parent == cur:
            break
        cur = cur.parent
    return start_dir.resolve()


def resolve_paths(
    base_dir: Optional[Path] = None,
    input_filename: str = DEFAULT_INPUT_FILENAME,
    output_filename: str = DEFAULT_OUTPUT_FILENAME,
) -> CADPaths:
    """Resolve input and output paths relative to this script.
    
    Inputs:
    - base_dir: directory containing this script (typically CAD/src).
    
    Outputs:
    - CADPaths with absolute paths to the input workbook, output workbook, and log directory.
    
    Errors:
    - None. This function does not touch the filesystem beyond path construction.
    """

    script_dir = Path(__file__).resolve().parent
    if base_dir is None:
        base_dir = _find_project_root(script_dir)

    base_dir = base_dir.resolve()
    inputs_dir = base_dir / "inputs"
    outputs_dir = base_dir / "outputs"

    input_path = inputs_dir / input_filename
    output_path = outputs_dir / output_filename

    log_path = outputs_dir / "cad_processor.log"
    execution_log_path = outputs_dir / "cad_processor_execution.log"

    return CADPaths(
        base_dir=base_dir,
        inputs_dir=inputs_dir,
        outputs_dir=outputs_dir,
        input_path=input_path,
        output_path=output_path,
        log_path=log_path,
        execution_log_path=execution_log_path,
    )


def _ensure_dirs(paths: CADPaths) -> None:
    paths.outputs_dir.mkdir(parents=True, exist_ok=True)
    paths.inputs_dir.mkdir(parents=True, exist_ok=True)


def _setup_logging(paths: CADPaths, verbose: bool = False) -> None:
    _ensure_dirs(paths)
    level = logging.DEBUG if verbose else logging.INFO
    fmt = "%(asctime)s | %(levelname)s | %(message)s"
    logging.basicConfig(
        level=level,
        format=fmt,
        handlers=[
            logging.FileHandler(paths.log_path, mode="a", encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )


def _sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(chunk_size), b""):
            h.update(chunk)
    return h.hexdigest()


# -------------------------
# Data quality accounting
# -------------------------

@dataclass
class DataQuality:
    input_sheet: str = ""
    header_row_1_based: int = 0
    total_rows_seen: int = 0

    rows_dropped_missing_keys: int = 0
    rows_dropped_missing_year: int = 0
    rows_dropped_summary_rows: int = 0
    rows_with_missing_cost: int = 0
    rows_with_missing_students: int = 0
    rows_with_negative_students: int = 0

    subject_year_groups: int = 0
    school_year_groups: int = 0
    undefined_cost_per_student_groups: int = 0
    no_activity_groups: int = 0

    def as_rows(self) -> List[Tuple[str, object]]:
        return [
            ("Input sheet", self.input_sheet),
            ("Header row (1-based)", self.header_row_1_based),
            ("Rows seen (including dropped rows)", self.total_rows_seen),
            ("Rows dropped (missing School/Subject fields)", self.rows_dropped_missing_keys),
            ("Rows dropped (year not detected)", self.rows_dropped_missing_year),
            ("Rows dropped (summary rows: Total/Sum/Result)", self.rows_dropped_summary_rows),
            ("Rows with missing cost values (treated as 0.0 for sums)", self.rows_with_missing_cost),
            ("Rows with missing student counts (treated as 0 for sums)", self.rows_with_missing_students),
            ("Rows with negative student counts (dropped)", self.rows_with_negative_students),
            ("Subject-Year groups", self.subject_year_groups),
            ("School-Year groups", self.school_year_groups),
            ("Groups with undefined cost per student (cost>0 and students==0)", self.undefined_cost_per_student_groups),
            ("Groups with no activity (cost==0 and students==0)", self.no_activity_groups),
        ]


# -------------------------
# Input table detection and parsing
# -------------------------

def _canonicalise_header_cell(value: object) -> str:
    if value is None:
        return ""
    norm = _normalise_header(str(value))
    return COLUMN_ALIASES.get(norm, str(value).strip())


def _extract_year(value: object) -> Optional[int]:
    if value is None:
        return None
    m = RE_YEAR.search(str(value))
    if not m:
        return None
    return int(m.group(0))


def _to_float(value: object) -> Tuple[float, bool]:
    """
    Convert to float. Returns (float_value, was_missing).

    Missing values are treated as 0.0 for aggregation, but counted via DataQuality.
    """
    if value is None:
        return 0.0, True
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value), False
    s = str(value).strip()
    if s == "":
        return 0.0, True
    # Remove currency symbols and grouping separators
    s = s.replace("$", "").replace(",", "")
    try:
        v = float(s)
        if not math.isfinite(v):
            return 0.0, True
        return v, False
    except ValueError:
        return 0.0, True


def _to_int(value: object) -> Tuple[int, bool]:
    """
    Convert to int. Returns (int_value, was_missing).

    Missing values are treated as 0 for aggregation, but counted via DataQuality.
    """
    if value is None:
        return 0, True
    if isinstance(value, (int, np.integer)):
        return int(value), False
    s = str(value).strip()
    if s == "":
        return 0, True
    s = s.replace(",", "")
    try:
        return int(float(s)), False
    except ValueError:
        return 0, True


@dataclass(frozen=True)
class LocatedTable:
    sheet_name: str
    header_row_1_based: int
    col_index: Dict[str, int]  # canonical col name -> 0-based index in row tuples


def locate_table(input_path: Path, scan_rows: int = 30, scan_cols: int = 60) -> LocatedTable:
    """
    Locate the first sheet and header row that contain all required columns.

    Deterministic search order: workbook sheet order, then header row from 1..scan_rows.
    """
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=scan_rows, max_col=scan_cols, values_only=True), start=1):
                headers = [_canonicalise_header_cell(v) for v in row]
                norm_headers = [_normalise_header(h) for h in headers]
                idx_map: Dict[str, int] = {}
                for i, h_norm in enumerate(norm_headers):
                    if h_norm in {_normalise_header(c) for c in REQUIRED_COLUMNS}:
                        canonical = COLUMN_ALIASES.get(h_norm, headers[i])
                        idx_map[canonical] = i
                # Verify all required columns present
                if all(c in idx_map for c in REQUIRED_COLUMNS):
                    logging.info("Detected input table: sheet='%s', header_row=%d", sheet, r_idx)
                    return LocatedTable(sheet_name=sheet, header_row_1_based=r_idx, col_index=idx_map)
    finally:
        wb.close()

    raise ValueError(
        "Could not locate a table containing the required columns: " + ", ".join(REQUIRED_COLUMNS)
    )


def stream_rows(input_path: Path, loc: LocatedTable, scan_cols: int = 60) -> Iterable[Tuple[object, ...]]:
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    ws = wb[loc.sheet_name]
    try:
        for row in ws.iter_rows(min_row=loc.header_row_1_based + 1, max_col=scan_cols, values_only=True):
            yield row
    finally:
        wb.close()


# -------------------------
# Aggregation
# -------------------------

@dataclass
class AggValue:
    cost_sum: float = 0.0
    students_sum: int = 0

    def add(self, cost: float, students: int) -> None:
        self.cost_sum += float(cost)
        self.students_sum += int(students)


def _safe_cost_per_student(total_cost: float, total_students: int) -> Tuple[Optional[float], str]:
    """
    Return (cost_per_student, status).

    Status values:
    - "OK": students>0
    - "No activity": cost==0 and students==0
    - "Undefined (0 students)": cost>0 and students==0
    """
    if total_students > 0:
        return float(total_cost) / float(total_students), "OK"
    if float(total_cost) == 0.0:
        return 0.0, "No activity"
    return None, "Undefined (0 students)"


def aggregate(input_path: Path) -> Tuple[pd.DataFrame, pd.DataFrame, DataQuality, Dict[int, "FuzzyAnchors"]]:
    """
    Returns (subject_year_df, school_year_df, dq, anchors_by_year).
    """
    loc = locate_table(input_path)
    dq = DataQuality(input_sheet=loc.sheet_name, header_row_1_based=loc.header_row_1_based)

    subject_year: Dict[Tuple[str, str, str, int], AggValue] = {}
    school_year: Dict[Tuple[str, int], AggValue] = {}

    for row in stream_rows(input_path, loc):
        dq.total_rows_seen += 1

        def _get(col: str) -> object:
            idx = loc.col_index[col]
            return row[idx] if idx < len(row) else None

        school = _get(COL_SCHOOL)
        subject_no = _get(COL_SUBJECT_NO)
        subject = _get(COL_SUBJECT)
        session = _get(COL_TEACHING_SESSION)

        # Drop summary rows (common in exported pivot outputs)
        if isinstance(school, str) and RE_SUMMARY_ROW.match(school.strip() or ""):
            dq.rows_dropped_summary_rows += 1
            continue
        if isinstance(subject, str) and RE_SUMMARY_ROW.match(subject.strip() or ""):
            dq.rows_dropped_summary_rows += 1
            continue

        if school is None or subject_no is None or subject is None:
            dq.rows_dropped_missing_keys += 1
            continue

        school_s = str(school).strip()
        subject_no_s = str(subject_no).strip()
        subject_s = str(subject).strip()

        if school_s == "" or subject_no_s == "" or subject_s == "":
            dq.rows_dropped_missing_keys += 1
            continue

        year = _extract_year(session)
        if year is None:
            dq.rows_dropped_missing_year += 1
            continue

        cost_raw = _get(COL_COST)
        students_raw = _get(COL_STUDENT_COUNT)

        cost, missing_cost = _to_float(cost_raw)
        students, missing_students = _to_int(students_raw)

        if missing_cost:
            dq.rows_with_missing_cost += 1
        if missing_students:
            dq.rows_with_missing_students += 1

        if students < 0:
            dq.rows_with_negative_students += 1
            continue

        skey = (school_s, subject_no_s, subject_s, int(year))
        if skey not in subject_year:
            subject_year[skey] = AggValue()
        subject_year[skey].add(cost, students)

        ykey = (school_s, int(year))
        if ykey not in school_year:
            school_year[ykey] = AggValue()
        school_year[ykey].add(cost, students)

    # Build Subject-Year dataframe
    subject_records: List[Dict[str, object]] = []
    for (school, subject_no, subject, year), agg in sorted(subject_year.items(), key=lambda t: t[0]):
        cps, status = _safe_cost_per_student(agg.cost_sum, agg.students_sum)
        subject_records.append(
            {
                COL_SCHOOL: school,
                COL_SUBJECT_NO: subject_no,
                COL_SUBJECT: subject,
                "Year": year,
                "Total_Oncosts": float(agg.cost_sum),
                "Total_Students": int(agg.students_sum),
                "Oncosts_Per_Student": cps,
                "Status": status,
            }
        )

    subject_df = pd.DataFrame.from_records(subject_records)
    dq.subject_year_groups = int(subject_df.shape[0])

    # Build School-Year dataframe
    school_records: List[Dict[str, object]] = []
    for (school, year), agg in sorted(school_year.items(), key=lambda t: t[0]):
        cps, status = _safe_cost_per_student(agg.cost_sum, agg.students_sum)
        school_records.append(
            {
                COL_SCHOOL: school,
                "Year": year,
                "Total_Oncosts": float(agg.cost_sum),
                "Total_Students": int(agg.students_sum),
                "Oncosts_Per_Student": cps,
                "Status": status,
            }
        )

    school_df = pd.DataFrame.from_records(school_records)
    dq.school_year_groups = int(school_df.shape[0])
    dq.undefined_cost_per_student_groups = int((subject_df["Status"] == "Undefined (0 students)").sum())
    dq.no_activity_groups = int((subject_df["Status"] == "No activity").sum())

    anchors_by_year = compute_year_anchors(school_df)

    return subject_df, school_df, dq, anchors_by_year


# -------------------------
# Report shaping
# -------------------------

def build_wide_report(subject_year: pd.DataFrame) -> pd.DataFrame:
    years = sorted(subject_year["Year"].unique().tolist())
    idx_cols = [COL_SCHOOL, COL_SUBJECT_NO, COL_SUBJECT]

    base = subject_year[idx_cols].drop_duplicates().sort_values(idx_cols).reset_index(drop=True)

    def _pivot(metric_col: str, col_suffix: str) -> pd.DataFrame:
        p = subject_year.pivot_table(
            index=idx_cols,
            columns="Year",
            values=metric_col,
            aggfunc="first",
            dropna=False,
        ).reindex(columns=years)
        p.columns = [f"{y} {col_suffix}" for y in p.columns]
        return p.reset_index()

    p_cost = _pivot("Total_Oncosts", "Incl Oncosts")
    p_cps = _pivot("Oncosts_Per_Student", "Incl Oncosts Per Student")
    p_students = _pivot("Total_Students", "Student Count")

    report = (
        base.merge(p_cost, on=idx_cols, how="left")
        .merge(p_cps, on=idx_cols, how="left")
        .merge(p_students, on=idx_cols, how="left")
    )

    # Fill totals for readability
    for y in years:
        report[f"{y} Incl Oncosts"] = pd.to_numeric(report[f"{y} Incl Oncosts"], errors="coerce").fillna(0.0)
        report[f"{y} Student Count"] = pd.to_numeric(report[f"{y} Student Count"], errors="coerce").fillna(0).astype(int)

    return report


def build_trend_matrix(school_year: pd.DataFrame) -> Tuple[pd.DataFrame, List[int]]:
    years = sorted(school_year["Year"].unique().tolist())
    pivot = school_year.pivot_table(
        index=COL_SCHOOL,
        columns="Year",
        values="Oncosts_Per_Student",
        aggfunc="first",
        dropna=False,
    ).reindex(columns=years)
    pivot = pivot.sort_index()
    return pivot, years


# -------------------------
# Fuzzy banding
# -------------------------

@dataclass(frozen=True)
class FuzzyAnchors:
    a_min: float
    b_median: float
    c_max: float


def _finite_positive(values: Iterable[Optional[float]]) -> List[float]:
    out: List[float] = []
    for v in values:
        if v is None:
            continue
        if isinstance(v, (float, int)) and math.isfinite(float(v)) and float(v) > 0.0:
            out.append(float(v))
    return out


def compute_year_anchors(school_year: pd.DataFrame) -> Dict[int, Optional[FuzzyAnchors]]:
    anchors: Dict[int, Optional[FuzzyAnchors]] = {}
    for year, g in school_year.groupby("Year", sort=True):
        vals = _finite_positive(g["Oncosts_Per_Student"].tolist())
        if not vals:
            anchors[int(year)] = None
            continue
        vals_sorted = sorted(vals)
        a = float(vals_sorted[0])
        c = float(vals_sorted[-1])
        b = float(np.median(vals_sorted))
        anchors[int(year)] = FuzzyAnchors(a_min=a, b_median=b, c_max=c)
    return anchors


def _mu_low(x: float, a: float, b: float) -> float:
    if b <= a:
        return 1.0 if x <= a else 0.0
    if x <= a:
        return 1.0
    if x >= b:
        return 0.0
    return (b - x) / (b - a)


def _mu_high(x: float, b: float, c: float) -> float:
    if c <= b:
        return 1.0 if x >= c else 0.0
    if x <= b:
        return 0.0
    if x >= c:
        return 1.0
    return (x - b) / (c - b)


def _mu_medium(x: float, a: float, b: float, c: float) -> float:
    if b <= a or c <= b or c <= a:
        return 1.0 if x == b else 0.0
    if x <= a or x >= c:
        return 0.0
    if x == b:
        return 1.0
    if x < b:
        return (x - a) / (b - a)
    return (c - x) / (c - b)


def build_fuzzy_bands(school_year: pd.DataFrame, anchors_by_year: Dict[int, Optional[FuzzyAnchors]]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    records: List[Dict[str, object]] = []

    for _, r in school_year.sort_values([COL_SCHOOL, "Year"]).iterrows():
        year = int(r["Year"])
        x = r["Oncosts_Per_Student"]
        if x is None or not isinstance(x, (float, int)) or not math.isfinite(float(x)) or float(x) <= 0.0:
            continue
        anchors = anchors_by_year.get(year)
        if anchors is None:
            continue

        a, b, c = anchors.a_min, anchors.b_median, anchors.c_max
        x = float(x)

        mu_l = _mu_low(x, a=a, b=b)
        mu_m = _mu_medium(x, a=a, b=b, c=c)
        mu_h = _mu_high(x, b=b, c=c)

        # Deterministic tie-breaking: Medium > Low > High
        band_order = [("Medium", mu_m), ("Low", mu_l), ("High", mu_h)]
        band = max(band_order, key=lambda t: (t[1], {"Medium": 2, "Low": 1, "High": 0}[t[0]]))[0]
        band_score = {"Low": 0.0, "Medium": 0.5, "High": 1.0}[band]

        records.append(
            {
                "School": r[COL_SCHOOL],
                "Year": year,
                "Cost per Student": x,
                "Mu_Low": round(mu_l, 4),
                "Mu_Medium": round(mu_m, 4),
                "Mu_High": round(mu_h, 4),
                "Band": band,
                "Band_Score_0_to_1": band_score,
                "Anchor_Min": round(a, 4),
                "Anchor_Median": round(b, 4),
                "Anchor_Max": round(c, 4),
            }
        )

    bands_df = pd.DataFrame.from_records(records)
    if bands_df.empty:
        summary = pd.DataFrame(columns=["Year", "Low", "Medium", "High", "Total"])
        return bands_df, summary

    summary = (
        bands_df.groupby(["Year", "Band"], sort=True)
        .size()
        .unstack(fill_value=0)
        .reindex(columns=["Low", "Medium", "High"], fill_value=0)
    )
    summary["Total"] = summary.sum(axis=1)
    summary = summary.reset_index()
    return bands_df, summary


# -------------------------
# Excel writing helpers
# -------------------------

def _set_workbook_properties(wb: openpyxl.Workbook) -> None:
    props = wb.properties
    props.creator = "CADProcessorFuzzy"
    props.lastModifiedBy = "CADProcessorFuzzy"
    fixed_dt = datetime.datetime(2000, 1, 1, 0, 0, 0)
    props.created = fixed_dt
    props.modified = fixed_dt
    props.title = "Processed CAD Contract"
    props.subject = "Deterministic preprocessing with fuzzy banding"


def _write_dataframe(ws: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    header_font = Font(bold=True)
    for j, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=j, value=str(col))
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for i in range(df.shape[0]):
        for j, col in enumerate(df.columns, start=start_col):
            val = df.iloc[i, j - start_col]
            if isinstance(val, float) and (math.isnan(val) or not math.isfinite(val)):
                val = None
            ws.cell(row=start_row + 1 + i, column=j, value=val)


def _autosize_columns(ws: openpyxl.worksheet.worksheet.Worksheet, max_width: int = 45) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), values_only=True):
        for idx, v in enumerate(row, start=1):
            if v is None:
                continue
            widths[idx] = max(widths.get(idx, 0), len(str(v)))
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(10, w + 2), max_width)


def _apply_number_formats_report(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for c, h in enumerate(headers, start=1):
        if h is None:
            continue
        h = str(h)
        if "Incl Oncosts Per Student" in h:
            fmt = "0.0000"
        elif "Incl Oncosts" in h:
            fmt = "0.00"
        elif "Student Count" in h:
            fmt = "0"
        else:
            continue
        for r in range(2, ws.max_row + 1):
            ws.cell(r, c).number_format = fmt


def _apply_number_formats_trend(ws: openpyxl.worksheet.worksheet.Worksheet, year_cols: List[int]) -> None:
    for c in year_cols:
        for r in range(2, ws.max_row + 1):
            ws.cell(r, c).number_format = "0.0000"


def _apply_number_formats_fuzzy(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    hdr_map = {str(h): i for i, h in enumerate(headers, start=1) if h is not None}
    for name in ["Cost per Student", "Mu_Low", "Mu_Medium", "Mu_High", "Anchor_Min", "Anchor_Median", "Anchor_Max"]:
        c = hdr_map.get(name)
        if c:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = "0.0000"
    c = hdr_map.get("Band_Score_0_to_1")
    if c:
        for r in range(2, ws.max_row + 1):
            ws.cell(r, c).number_format = "0.0"


# -------------------------
# Sheet writers
# -------------------------

def write_processing_summary(ws: openpyxl.worksheet.worksheet.Worksheet, dq: DataQuality, input_hash: str, anchors: Dict[int, Optional[FuzzyAnchors]]) -> None:
    """Write the Processing_Summary worksheet.
    
    Purpose:
    - Provide a compact, auditable run record derived from observed inputs and processing counts.
    
    Contents:
    - Input file identity (name, path, size, SHA-256)
    - Detected sheet and header row
    - Row and group counts produced during processing
    - Per-year anchors used for fuzzy banding
    """

    ws["A1"] = "Processing Summary"
    ws["A1"].font = Font(bold=True, size=14)

    rows = [("Input file SHA-256", input_hash)] + dq.as_rows()
    r = 3
    for k, v in rows:
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
        r += 1

    # Anchor summary (deterministic, computed)
    ws.cell(r + 1, 1, "Anchors used (finite positive School-Year values)").font = Font(bold=True)
    rr = r + 2
    for year in sorted(anchors.keys()):
        an = anchors[year]
        ws.cell(rr, 1, str(year)).font = Font(bold=True)
        if an is None:
            ws.cell(rr, 2, "No finite positive values")
        else:
            ws.cell(rr, 2, f"min {an.a_min:.4f}, median {an.b_median:.4f}, max {an.c_max:.4f}")
        rr += 1

    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 70
    ws.freeze_panes = "A3"


def write_trend_analysis(
    wb: openpyxl.Workbook,
    trend: pd.DataFrame,
    years: List[int],
    school_year: pd.DataFrame,
    anchors_by_year: Dict[int, Optional[FuzzyAnchors]],
) -> None:
    """Write the Trend_Analysis worksheet.
    
    Contents:
    - School-Year cost-per-student table with conditional formatting and a concise methodology block.
    
    Notes:
    - Any narrative text is descriptive of the implemented computation and avoids unsupported claims.
    """

    ws = wb.create_sheet("Trend Analysis")

    df_out = trend.copy()
    df_out.insert(0, "School", df_out.index)
    df_out = df_out.reset_index(drop=True)

    _write_dataframe(ws, df_out)
    ws.freeze_panes = "B2"

    year_cols = list(range(2, 2 + len(years)))
    _apply_number_formats_trend(ws, year_cols)

    # Conditional formatting (green -> yellow -> red) per year column using numeric anchors
    for i, year in enumerate(years):
        col = 2 + i
        anchors = anchors_by_year.get(int(year))
        if anchors is None:
            continue
        a, b, c = anchors.a_min, anchors.b_median, anchors.c_max
        if not (c > a):
            continue
        rng = f"{get_column_letter(col)}2:{get_column_letter(col)}{ws.max_row}"
        rule = ColorScaleRule(
            start_type="num", start_value=a, start_color="FF63BE7B",
            mid_type="num", mid_value=b, mid_color="FFFFEB84",
            end_type="num", end_value=c, end_color="FFF8696B",
        )
        ws.conditional_formatting.add(rng, rule)

    # Methodology & Insights block (computed and therefore auditable)
    block_col = 7  # G
    ws.cell(2, block_col, "Methodology & Insights").font = Font(bold=True, size=12)
    ws.cell(2, block_col).alignment = Alignment(wrap_text=True, vertical="top")

    status_counts = school_year["Status"].value_counts().to_dict()
    no_activity = int(status_counts.get("No activity", 0))
    undefined = int(status_counts.get("Undefined (0 students)", 0))

    anchor_lines: List[str] = []
    for y in years:
        an = anchors_by_year.get(int(y))
        if an is None:
            anchor_lines.append(f"{y}: (no finite positive values)")
        else:
            anchor_lines.append(f"{y}: min {an.a_min:.2f}, median {an.b_median:.2f}, max {an.c_max:.2f}")

    lines = [
        "1. Aggregation: For each School and Year, Total_Oncosts and Total_Students are aggregated across subjects. "
        "Cost per student is Total_Oncosts / Total_Students when Total_Students > 0.",
        f"2. Zero and undefined cases: {no_activity} School-Year cells indicate no recorded activity (oncosts=0 and students=0). "
        f"{undefined} School-Year cells are undefined due to zero student count with non-zero oncosts and are left blank.",
        "3. Scaling: Heatmap anchors are computed from finite, positive cost-per-student values per year (min/median/max): "
        + "; ".join(anchor_lines) + ".",
        "4. Fuzzy bands: The same anchors define simple triangular membership functions for Low, Medium, and High bands "
        "in the 'Fuzzy Bands' sheet.",
    ]

    r = 4
    for ln in lines:
        ws.cell(r, block_col, ln).alignment = Alignment(wrap_text=True, vertical="top")
        r += 2

    _autosize_columns(ws)


def write_report_sheet(wb: openpyxl.Workbook, report_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Report")
    _write_dataframe(ws, report_df)
    ws.freeze_panes = "A2"
    _apply_number_formats_report(ws)
    _autosize_columns(ws)


def write_fuzzy_bands_sheet(wb: openpyxl.Workbook, bands_df: pd.DataFrame, summary_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Fuzzy Bands")

    if bands_df.empty:
        ws["A1"] = "No finite, positive School-Year cost-per-student values available for fuzzy banding."
        ws["A1"].alignment = Alignment(wrap_text=True)
        return

    _write_dataframe(ws, bands_df)
    _apply_number_formats_fuzzy(ws)
    ws.freeze_panes = "A2"

    start_row = bands_df.shape[0] + 4
    ws.cell(start_row, 1, "Summary (counts)").font = Font(bold=True)
    start_row += 1
    _write_dataframe(ws, summary_df, start_row=start_row, start_col=1)

    _autosize_columns(ws)


# -------------------------
# Main processor
# -------------------------

class CADProcessorFuzzy:
    def __init__(self, paths: CADPaths, write_processing_summary_sheet: bool = True) -> None:
        self.paths = paths
        self.write_processing_summary_sheet = bool(write_processing_summary_sheet)

    def run(self) -> Path:
        if not self.paths.input_path.exists():
            tried = [
                str(self.paths.input_path),
                str(Path(__file__).resolve().parent / "inputs" / self.paths.input_path.name),
                str(Path.cwd() / "inputs" / self.paths.input_path.name),
            ]
            msg = (
                "Input file not found.\n"
                f"Expected at: {self.paths.input_path}\n"
                "Also checked:\n- " + "\n- ".join(tried[1:]) + "\n"
                f"Project root resolved as: {self.paths.base_dir}\n"
                "Fix: place the input workbook in '<project_root>/inputs/' or pass --input-filename."
            )
            raise FileNotFoundError(msg)

        logging.info("Input: %s", self.paths.input_path)
        input_hash = _sha256_file(self.paths.input_path)
        logging.info("Input SHA-256: %s", input_hash)

        subject_df, school_df, dq, anchors_by_year = aggregate(self.paths.input_path)

        report_df = build_wide_report(subject_df)
        trend_matrix, years = build_trend_matrix(school_df)
        bands_df, summary_df = build_fuzzy_bands(school_df, anchors_by_year)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _set_workbook_properties(wb)

        # Deterministic sheet order
        if self.write_processing_summary_sheet:
            ws_summary = wb.create_sheet("Processing Summary")
            write_processing_summary(ws_summary, dq, input_hash, anchors_by_year)

        write_trend_analysis(wb, trend_matrix, years, school_df, anchors_by_year)
        write_report_sheet(wb, report_df)
        write_fuzzy_bands_sheet(wb, bands_df, summary_df)

        self.paths.outputs_dir.mkdir(parents=True, exist_ok=True)
        wb.save(self.paths.output_path)
        logging.info("Output written: %s", self.paths.output_path)

        with self.paths.execution_log_path.open("a", encoding="utf-8") as f:
            f.write(f"OK\t{self.paths.input_path.name}\t{self.paths.output_path.name}\t{input_hash}\n")

        return self.paths.output_path


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Deterministic CAD contract processor with fuzzy banding.")
    p.add_argument("--base-dir", type=str, default=None, help="Project root containing inputs/, outputs/, config.json.")
    p.add_argument("--input-filename", type=str, default=DEFAULT_INPUT_FILENAME, help="Input file name in inputs/.")
    p.add_argument("--output-filename", type=str, default=DEFAULT_OUTPUT_FILENAME, help="Output file name in outputs/.")
    p.add_argument("--verbose", action="store_true", help="Verbose logging.")
    p.add_argument("--no-processing-summary", action="store_true", help="Do not add the Processing Summary sheet.")
    return p.parse_args()


def main() -> None:
    """CLI entry point.
    
    Behaviour:
    - Resolves paths, runs the processor, and exits non-zero on unhandled errors.
    """

    args = parse_args()
    base_dir = Path(args.base_dir) if args.base_dir else None
    paths = resolve_paths(base_dir=base_dir, input_filename=args.input_filename, output_filename=args.output_filename)
    _setup_logging(paths, verbose=bool(args.verbose))

    CADProcessorFuzzy(paths, write_processing_summary_sheet=(not args.no_processing_summary)).run()


if __name__ == "__main__":
    main()
